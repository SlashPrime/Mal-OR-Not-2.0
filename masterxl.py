import openpyxl
from openpyxl import Workbook

book = Workbook()

book.create_sheet("File")
book.create_sheet("URL")
book.create_sheet("IP")
book.create_sheet("E-mail")
book.create_sheet("Domain")
book.create_sheet("Phone no.")

book.active = book['File']
sheet=book.active
filerow1 = ['File','Malicious','Suspicious','Harmless','Undetected']
sheet.append(filerow1)
sheet.column_dimensions['A'].width = 15
sheet.column_dimensions['B'].width = 20
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 20

book.active = book['URL']
sheet=book.active
urlrow1 = ['URL','Harmless','Malicious','Suspicious','Undetected']
sheet.append(urlrow1)
sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 15

book.active = book['IP']
sheet=book.active
iprow1 = ['IP Address','Fraud Score','Proxy','Vpn','Tor','Country name','Region name','City name','Latitude','Longitude','Zip code','Time zone','Isp','Domain','Geographical Location']
sheet.append(iprow1)
sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 10
sheet.column_dimensions['D'].width = 10
sheet.column_dimensions['E'].width = 10
sheet.column_dimensions['F'].width = 15
sheet.column_dimensions['G'].width = 15
sheet.column_dimensions['H'].width = 15
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 10
sheet.column_dimensions['M'].width = 35
sheet.column_dimensions['N'].width = 20
sheet.column_dimensions['O'].width = 80

book.active = book['E-mail']
sheet=book.active
emailrow1 = ['E-mail Address','Valid','Disposable','Smtp score','dns valid','Honeypot','Deliverability','Recent abuse','Fraud score','Leaked']
sheet.append(emailrow1)
sheet.column_dimensions['A'].width = 25
sheet.column_dimensions['B'].width = 10
sheet.column_dimensions['C'].width = 10
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 10
sheet.column_dimensions['F'].width = 10
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12
sheet.column_dimensions['I'].width = 12
sheet.column_dimensions['J'].width = 10

book.active = book['Domain']
sheet=book.active
domainrow1 = ['Domain name', 'Registrar', 'Updated date', 'Creation date', 'Expiration date', 'Country', 'Unsafe', 'Ip address', 'Server', 'Status code', 'Domain rank', 'Spamming', 'Malware', 'Phishing', 'Suspicious', 'Adult', 'Risk score', 'Category']
sheet.append(domainrow1)
sheet.column_dimensions['A'].width = 15
sheet.column_dimensions['B'].width = 22
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 20
sheet.column_dimensions['F'].width = 10
sheet.column_dimensions['G'].width = 10
sheet.column_dimensions['H'].width = 17
sheet.column_dimensions['I'].width = 15
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 10
sheet.column_dimensions['M'].width = 10
sheet.column_dimensions['N'].width = 10
sheet.column_dimensions['O'].width = 10
sheet.column_dimensions['P'].width = 10
sheet.column_dimensions['Q'].width = 10
sheet.column_dimensions['R'].width = 15

book.active = book['Phone no.']
sheet=book.active
phnorow1 = ['Number','Valid','Local format','International format','Country prefix','Country code','Country name','Location','Carrier','Line type']
sheet.append(phnorow1)
sheet.column_dimensions['A'].width = 15
sheet.column_dimensions['B'].width = 10
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 15
sheet.column_dimensions['F'].width = 15
sheet.column_dimensions['G'].width = 20
sheet.column_dimensions['H'].width = 15
sheet.column_dimensions['I'].width = 30
sheet.column_dimensions['J'].width = 12

del book['Sheet']

book.save("masterexcel.xlsx")
